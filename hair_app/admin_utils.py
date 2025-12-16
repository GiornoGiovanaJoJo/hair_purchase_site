"""
Утилиты для экспорта данных из админки
"""
import csv
from io import BytesIO, StringIO
from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def export_applications_to_csv(queryset):
    """
    Экспортировать заявки в CSV
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="applications.csv"'
    
    writer = csv.writer(response, delimiter=';')
    
    # Заголовки
    writer.writerow([
        'ID',
        'Имя',
        'Телефон',
        'Email',
        'Город',
        'Длина волос',
        'Цвет',
        'Структура',
        'Состояние',
        'Возраст',
        'Смета',
        'Финальная цена',
        'Статус',
        'Дата создания',
    ])
    
    # Данные
    for app in queryset:
        writer.writerow([
            app.id,
            app.name,
            app.phone,
            app.email,
            app.city or '-',
            app.length,
            app.color,
            app.structure,
            app.condition,
            app.age,
            app.estimated_price or '-',
            app.final_price or '-',
            dict(app.STATUS_CHOICES).get(app.status, app.status),
            app.created_at.strftime('%d.%m.%Y %H:%M'),
        ])
    
    return response


def export_applications_to_excel(queryset):
    """
    Экспортировать заявки в Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Заявки'
    
    # Стили
    header_fill = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовки
    headers = [
        'ID',
        'Имя',
        'Телефон',
        'Email',
        'Город',
        'Длина',
        'Цвет',
        'Структура',
        'Состояние',
        'Возраст',
        'Смета (RUB)',
        'Финал (RUB)',
        'Статус',
        'Дата',
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Данные
    for row_idx, app in enumerate(queryset, 2):
        row_data = [
            app.id,
            app.name,
            app.phone,
            app.email,
            app.city or '-',
            app.length,
            app.color,
            app.structure,
            app.condition,
            app.age,
            app.estimated_price or '-',
            app.final_price or '-',
            dict(app.STATUS_CHOICES).get(app.status, app.status),
            app.created_at.strftime('%d.%m.%Y %H:%M'),
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Форматирование чисел
            if col in [11, 12]:  # Цены
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
    
    # Ширина колонок
    widths = [8, 15, 18, 20, 12, 12, 15, 15, 15, 12, 14, 14, 12, 18]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    # Сохранение
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="applications.xlsx"'
    wb.save(response)
    return response


def export_prices_to_excel(queryset):
    """
    Экспортировать цены в Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Прайс-лист'
    
    # Стили
    header_fill = PatternFill(start_color='0f3460', end_color='0f3460', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовки
    headers = ['ID', 'Длина', 'Цвет', 'Структура', 'Состояние', 'Цена (RUB)', 'Активна']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Данные
    for row_idx, price in enumerate(queryset, 2):
        row_data = [
            price.id,
            price.length,
            price.color,
            price.structure,
            price.condition,
            price.base_price,
            'Да' if price.is_active else 'Нет',
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            if col == 6:  # Цена
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right', vertical='center')
    
    # Ширина колонок
    widths = [8, 12, 15, 15, 15, 14, 10]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="pricelist.xlsx"'
    wb.save(response)
    return response
